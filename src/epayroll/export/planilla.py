from __future__ import annotations

from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any, BinaryIO

from fpdf import FPDF
from fpdf.enums import XPos, YPos
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _to_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _money(value: Any) -> str:
    dec = _to_decimal(value)
    if dec is None:
        return "—"
    return f"{dec:,.2f}"


def _cell_value(row: dict[str, Any], col: dict[str, Any]) -> Any:
    raw = row.get(col["key"])
    if col.get("tipo") == "moneda":
        dec = _to_decimal(raw)
        return float(dec) if dec is not None else None
    if col.get("tipo") == "numero":
        dec = _to_decimal(raw)
        return float(dec) if dec is not None else raw
    return raw if raw is not None else ""


def _safe_text(value: str) -> str:
    return value.encode("latin-1", "replace").decode("latin-1")


def generate_planilla_xlsx(data: dict[str, Any], output: Path | BinaryIO) -> None:
    """Genera hoja Excel de verificación planilla."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Planilla"

    periodo = data.get("periodo") or {}
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=11)
    header_fill = PatternFill("solid", fgColor="1A2E50")
    header_font_white = Font(bold=True, color="FFFFFF", size=9)
    total_fill = PatternFill("solid", fgColor="E8EEF4")

    ws["A1"] = data.get("razon_social") or "Planilla"
    ws["A1"].font = header_font
    ws["A2"] = (
        f"Período: {periodo.get('fecha_inicio', '')} → {periodo.get('fecha_fin', '')} · "
        f"Pago: {periodo.get('fecha_pago', '')} · Tipo: {periodo.get('tipo', '')}"
    )
    ws["A2"].font = Font(size=10, italic=True)

    cols = data.get("columnas") or []
    rows = data.get("rows") or []
    totales = data.get("totales") or {}
    start_row = 4

    for idx, col in enumerate(cols, start=1):
        cell = ws.cell(row=start_row, column=idx, value=col.get("titulo") or col.get("key"))
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    money_cols = {i + 1 for i, c in enumerate(cols) if c.get("tipo") == "moneda"}
    num_cols = {i + 1 for i, c in enumerate(cols) if c.get("tipo") == "numero"}

    for r_idx, row in enumerate(rows, start=start_row + 1):
        for c_idx, col in enumerate(cols, start=1):
            value = _cell_value(row, col)
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if c_idx in money_cols and isinstance(value, (int, float)):
                cell.number_format = '"B/." #,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif c_idx in num_cols and isinstance(value, (int, float)):
                cell.number_format = "0.##"
                cell.alignment = Alignment(horizontal="right")

    total_row = start_row + len(rows) + 1
    for c_idx, col in enumerate(cols, start=1):
        cell = ws.cell(row=total_row, column=c_idx)
        cell.fill = total_fill
        cell.font = title_font
        if c_idx == 1:
            cell.value = "TOTAL"
        elif col["key"] in totales:
            dec = _to_decimal(totales[col["key"]])
            cell.value = float(dec) if dec is not None else totales[col["key"]]
            if c_idx in money_cols and isinstance(cell.value, (int, float)):
                cell.number_format = '"B/." #,##0.00'
            cell.alignment = Alignment(horizontal="right")

    for idx, col in enumerate(cols, start=1):
        letter = get_column_letter(idx)
        title = col.get("titulo") or col.get("key") or ""
        width = max(10, min(28, len(title) + 2))
        if col.get("key") == "nombre_completo":
            width = 28
        ws.column_dimensions[letter].width = width

    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    wb.save(output)


def generate_planilla_pdf(data: dict[str, Any], output: Path | BinaryIO) -> None:
    """Genera PDF apaisado (A3) de verificación planilla."""
    cols = data.get("columnas") or []
    rows = data.get("rows") or []
    totales = data.get("totales") or {}
    periodo = data.get("periodo") or {}

    pdf = FPDF(orientation="L", unit="mm", format="A3")
    pdf.set_auto_page_break(auto=True, margin=8)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 8, _safe_text(data.get("razon_social") or "Verificación de planilla"), new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(
        0,
        6,
        _safe_text(
            f"Período {periodo.get('fecha_inicio', '')} → {periodo.get('fecha_fin', '')} · "
            f"Pago {periodo.get('fecha_pago', '')} · {len(rows)} empleado(s)"
        ),
        new_x=XPos.LMARGIN,
        new_y=YPos.NEXT,
    )
    pdf.ln(2)

    page_w = pdf.w - pdf.l_margin - pdf.r_margin
    sticky_keys = {"ficha", "nombre_completo", "cedula"}
    weights = []
    for col in cols:
        key = col.get("key", "")
        if key in sticky_keys:
            weights.append(2.2 if key == "nombre_completo" else 1.4)
        elif col.get("tipo") == "moneda":
            weights.append(1.1)
        else:
            weights.append(1.0)
    total_w = sum(weights)
    col_widths = [page_w * (w / total_w) for w in weights]
    line_h = 4.5
    font_size = 5.5 if len(cols) > 18 else 6.5

    def draw_row(values: list[str], *, header: bool = False, total: bool = False) -> None:
        if header:
            pdf.set_font("Helvetica", "B", font_size)
            pdf.set_fill_color(26, 46, 80)
            pdf.set_text_color(255, 255, 255)
        elif total:
            pdf.set_font("Helvetica", "B", font_size)
            pdf.set_fill_color(232, 238, 244)
            pdf.set_text_color(0, 0, 0)
        else:
            pdf.set_font("Helvetica", "", font_size)
            pdf.set_fill_color(255, 255, 255)
            pdf.set_text_color(0, 0, 0)

        for width, text in zip(col_widths, values):
            display = _safe_text(text)
            max_chars = max(4, int(width / 1.6))
            if len(display) > max_chars:
                display = display[: max_chars - 3] + "..."
            align = "R" if text and text.replace(",", "").replace(".", "").replace("-", "").replace("B", "").replace("/", "").strip().replace(" ", "").isdigit() else "L"
            pdf.cell(width, line_h, display, border=1, fill=header or total, align=align)
        pdf.ln()

    headers = [str(c.get("titulo") or c.get("key") or "") for c in cols]
    draw_row(headers, header=True)

    for row in rows:
        values = []
        for col in cols:
            raw = row.get(col["key"])
            if col.get("tipo") == "moneda":
                values.append(_money(raw))
            elif col.get("tipo") == "numero":
                dec = _to_decimal(raw)
                values.append(str(dec) if dec is not None else str(raw or ""))
            else:
                values.append(str(raw or ""))
        draw_row(values)

    total_values = []
    for i, col in enumerate(cols):
        if i == 0:
            total_values.append("TOTAL")
        elif col["key"] in totales:
            total_values.append(_money(totales[col["key"]]) if col.get("tipo") == "moneda" else str(totales[col["key"]]))
        else:
            total_values.append("")
    draw_row(total_values, total=True)

    pdf.set_y(-10)
    pdf.set_font("Helvetica", "I", 7)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 5, _safe_text("Verificación de planilla — EPayRoll"), align="C")

    if isinstance(output, Path):
        output.parent.mkdir(parents=True, exist_ok=True)
        pdf.output(str(output))
    else:
        pdf.output(output)


def planilla_export_filename(data: dict[str, Any], ext: str) -> str:
    periodo = data.get("periodo") or {}
    fin = periodo.get("fecha_fin") or "planilla"
    org = (data.get("razon_social") or "planilla").split()[0][:20]
    safe_org = "".join(ch if ch.isalnum() else "_" for ch in org)
    return f"planilla_{safe_org}_{fin}.{ext}"


def planilla_xlsx_bytes(data: dict[str, Any]) -> bytes:
    buf = BytesIO()
    generate_planilla_xlsx(data, buf)
    return buf.getvalue()


def planilla_pdf_bytes(data: dict[str, Any]) -> bytes:
    buf = BytesIO()
    generate_planilla_pdf(data, buf)
    return buf.getvalue()
