"""Exportación Excel según docs/Planilla_modelo.xlsx."""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font

from epayroll.db.legal_config_repository import PlanillaViewRepository


def _money(v: str | Decimal | None) -> float | str:
    if v is None or v == "":
        return ""
    return float(Decimal(str(v)))


def generate_planilla_modelo_xlsx(
    run_id: str,
    output_path: Path,
    database_url: str | None = None,
) -> dict[str, Any]:
    repo = PlanillaViewRepository()
    data = repo.get_run_planilla(run_id, database_url=database_url)
    wb = Workbook()
    ws = wb.active
    ws.title = "PLANILLA CSS"

    ws["B1"] = "1ERA QUINCENA" if data["periodo"]["tipo"] == "QUINCENAL" else data["periodo"]["tipo"]
    ws["B2"] = "PERIODO DE PAGO"
    ws["D2"] = f"{data['periodo']['fecha_inicio']} AL {data['periodo']['fecha_fin']}"
    ws["B3"] = "Planilla"
    ws["P3"] = "Descuentos"
    ws["Z3"] = "Gastos Empresa"

    col_keys = [c["key"] for c in data["columnas"]]
    headers = [c["titulo"] for c in data["columnas"]]
    for i, title in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=i, value=title)
        cell.font = Font(bold=True)

    for r_idx, row in enumerate(data["rows"], start=7):
        for c_idx, key in enumerate(col_keys, start=1):
            val = row.get(key, "")
            if key in {
                "salario_mensual", "salario_quincenal", "dias_pago", "monto_desc_dias",
                "dev_isr", "salario_cotizable", "css_empleado", "se_empleado", "isr",
                "cpp_prestaciones", "prestamo_empleado", "desc_prestamo", "descuento_banco",
                "saldo_prestamo", "total_descuentos", "cancelacion", "css_patronal",
                "se_patronal", "riesgo_profesional", "gastos_empresa", "total_cpp_prest",
            }:
                val = _money(val)
            ws.cell(row=r_idx, column=c_idx, value=val)

    total_row = 7 + len(data["rows"]) + 4
    ws.cell(row=total_row, column=2, value="TOTAL").font = Font(bold=True)
    for c_idx, key in enumerate(col_keys, start=1):
        if key in data.get("totales", {}):
            ws.cell(row=total_row, column=c_idx, value=_money(data["totales"][key]))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return {
        "run_id": run_id,
        "archivo": str(output_path),
        "filas": len(data["rows"]),
    }
